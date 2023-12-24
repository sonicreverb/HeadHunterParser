import os
import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink


def write_to_excel(filename, data):
    header_mapping = {
        'Должность': 'Post', 'Желаемая зарплата': 'Salary', 'Требуемый опыт работы': 'ReqWorkExp',
        'Тип занятости': 'TypeOfEmployment', 'Название компании': 'CompanyName', 'О Компании': 'CompanyInfo',
        'Основные задачи': 'Duties', 'Требования': 'Requirements', 'Мы предлагаем': 'Working conditions',
        'Пометка внизу': 'Downmark', 'Ключевые навыки': 'ReqSkills', 'Контактное лицо': 'ContactName',
        'Телефон': 'ContactPhone', 'Почта': 'ContactMail', 'Ссылка на вакансию': 'VacancyURL',
        'Сайт компании': 'CompanySiteURL', 'Сферы деятельности': 'CompanyActitvityAreas', 'Адрес': 'CompanyAddress',
        'Описание компании': 'CompanyDescription', 'Ссылка на нанимателя': 'CompanyHHPageURL'
    }

    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = list(header_mapping.keys())

        # Добавляем пустой столбец перед каждым столбцом с информацией
        for col_num, header in enumerate(headers, 1):
            # Создаем тонкий пустой столбец
            cell = ws.cell(row=1, column=col_num * 2, value=' ')
            cell.font = openpyxl.styles.Font(size=12)
            cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'))

            # Заголовок столбца с информацией
            cell = ws.cell(row=1, column=col_num * 2 + 1, value=header)
            cell.font = openpyxl.styles.Font(name='Times New Roman', size=14, bold=True)

    for row_num, row_data in enumerate(data, ws.max_row + 1):
        for col_num, header in enumerate(header_mapping.keys(), 1):
            key = header_mapping[header]
            # Записываем данные с учетом пустого столбца
            cell = ws.cell(row=row_num, column=col_num * 2, value=' ')
            cell = ws.cell(row=row_num, column=col_num * 2 + 1, value=row_data[key])
            if 'Ссылка' in header or 'Сайт' in header:
                cell.hyperlink = Hyperlink(row_data[key])
                cell.font = openpyxl.styles.Font(name='Times New Roman', size=12, underline='single', color='0000FF')
            else:
                cell.font = openpyxl.styles.Font(name='Times New Roman', size=12)

    # Автоматически подгоняем ширину столбцов под содержимое
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        if max_length > 50:
            max_length = 50
        adjusted_width = (max_length + 2)
        ws.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

    wb.save(filename)
    print(f'[XLSX CONNCETOR] Данные успешно записаны в файл \'{filename}\'.')
